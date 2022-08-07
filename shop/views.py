from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.views import APIView
from shop.models import *
from shop.serializers import ProductSerializer, ProductNameSerializer, BrandNameSerializer, BasketSerializer


class ProductNameView(generics.ListCreateAPIView):
    queryset = ProductName.objects.all()
    serializer_class = ProductNameSerializer


class BrandNameView(generics.ListCreateAPIView):
    queryset = BrandName.objects.all()
    serializer_class = BrandNameSerializer


class ProductsView(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer


class BasketView(generics.ListCreateAPIView):
    queryset = Basket.objects.all()
    serializer_class = BasketSerializer


class ExportBasketInExcel(APIView):
    def get(self, request):
        user_id = request.user.id
        basket = Basket.objects.filter(user_id=user_id)

        alignment = Alignment(horizontal='center', vertical='center', )
        workbook = Workbook()
        ws = workbook.active
        fields = ['User', 'Order']
        column = 1
        row = 1
        column_alf = ['A', 'B']
        column_weight = [10, 15]

        for field in fields:
            ws.column_dimensions[column_alf[column - 1]].width = column_weight[column - 1]
            cell = ws.cell(column=column, row=1, value=field)

            cell.alignment = alignment
            column += 1

        for product in basket:
            row += 1

            ws.cell(column=1, row=row, value=product.user_id).alignment = alignment
            ws.cell(column=2, row=row, value=product.order_id).alignment = alignment

        response = HttpResponse(content_type='application/ms_excel')
        response['Content_Disposition'] = f'attachment; filename = Data.xlsx'
        workbook.save(response)
        return response
